import tkinter as tk
from tkinter import filedialog
import openpyxl
import time
import pyautogui

tempo = 1
cood1 = float(input("Coordenadas do campo NUMERO: "))
cood2 = float(input("Coordenadas do documento encontrado: "))


def ler_planilha():
    arquivo_excel = filedialog.askopenfilename(filetypes=[('Arquivos Excel', '*.xlsx')])

    if arquivo_excel:
        try:
            wb = openpyxl.load_workbook(arquivo_excel)
            sheet = wb.active

            max_linha = sheet.max_row
            max_coluna = sheet.max_column

            janela = tk.Tk()
            janela.title('Valores da Planilha')

            for row in range(1, max_linha + 1):
                for col in range(1, max_coluna + 1):
                    valor = sheet.cell(row=row, column=col).value

                    if valor is None:
                        continue

                    if col % 2 == 0:
                        botao = tk.Button(janela, text=str(valor), command=lambda r=row, c=col: executar_acao(r, c, sheet))
                        botao.grid(row=row, column=col, padx=5, pady=5)
                    else:
                        botao = tk.Button(janela, text=str(valor), command=lambda r=row, c=col: executar_acao_impar(r, c, sheet))
                        botao.grid(row=row, column=col, padx=5, pady=5)

            janela.mainloop()

        except openpyxl.utils.exceptions.InvalidFileException:
            print('Arquivo inválido.')

def executar_acao(row, col, sheet):
    numero = sheet.cell(row=row, column=col).value

    time.sleep(1)
    pyautogui.click(780, 290)
    time.sleep(tempo)
    pyautogui.press('tab', presses=15)
    time.sleep(tempo)
    pyautogui.write("180")
    time.sleep(tempo)
    pyautogui.press('enter')
    time.sleep(tempo)
    pyautogui.write(str(numero))
    time.sleep(tempo)
    pyautogui.press('enter')
    pyautogui.keyDown('ctrl')
    pyautogui.press('enter')
    pyautogui.keyUp('ctrl')
        #data aqui
#    time.sleep(2)
#    pyautogui.click(600, 350)
#    pyautogui.click(600, 350)
#    time.sleep(tempo)
#    pyautogui.write("09082024")
#    pyautogui.press('enter')

def executar_acao_impar(row, col, sheet):
    numero = sheet.cell(row=row, column=col).value

    pyautogui.click('cood1')
    pyautogui.click('cood1')
    pyautogui.write(str(numero))
    pyautogui.press('f3')
    pyautogui.click(cood2)
    pyautogui.click(cood2)

janela_principal = tk.Tk()
janela_principal.title('Leitor de Planilha')

botao_selecionar = tk.Button(janela_principal, text='Selecionar Planilha', command=ler_planilha)
botao_selecionar.pack(padx=10, pady=10)

janela_principal.mainloop()